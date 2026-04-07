#!/usr/bin/env python3
"""发票分类与汇总表生成脚本 v4 - 智能去重版 - 高级OCR"""
import os, sys, argparse, re, subprocess
from datetime import datetime
from pathlib import Path
from collections import defaultdict

# 尝试导入高级OCR模块
try:
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'invoice-reader'))
    from advanced_ocr import extract_amount_from_text as ocr_extract_amount, extract_date as ocr_extract_date, extract_invoice_number as ocr_extract_invoice_number, classify as ocr_classify
    ADVANCED_OCR = True
except ImportError:
    ADVANCED_OCR = False

CATEGORIES = {
    "交通出行": ["交通", "出行", "打车", "滴滴", "出租车", "火车票", "机票", "航空", "地铁", "公交", "高速", "ETC", "汽油", "充电", "uber", "didi", "高德", "行程", "停车费", "过路费", "船票", "船运", "阳光出行"],
    "餐饮美食": ["餐饮", "餐费", "餐服务", "饭店", "餐厅", "午餐", "晚餐", "美食", "外卖", "奶茶", "咖啡", "酒楼"],
    "办公用品": ["办公", "文具", "打印", "耗材", "电脑", "设备", "配件", "办公用品", "纸张", "墨盒", "office", "半导体", "电子元件", "立创"],
    "酒店住宿": ["酒店", "住宿", "宾馆", "旅馆", "民宿", "房费", "hotel", "inn", "room"],
}

# 交通出行子分类（按优先级排序）
TRANSPORT_SUBCATEGORIES = {
    "机票": ["机票", "flight", "航空", "airline", "携程机票", "ia_rsv"],
    "火车票": ["火车票", "铁路", "动车", "高铁"],
    "船票": ["船票", "船运", "海天码头"],
    "停车费": ["停车费", "停车服务", "parking"],
    "ETC": ["etc", "高速", "过路费", "通行费", "收费公路"],
    "网约车": ["滴滴", "didi", "uber", "高德", "打车", "网约车", "行程", "itinerary", "阳光出行", "第三方网约车服务公司"],
    "汽油": ["汽油", "加油", "燃油", "中石化", "中石油", "壳牌"],
    "充电": ["充电", "电动车", "充电桩", "新能源"],
    "代驾": ["代驾", "代驾服务"],
    "租车": ["租车", "car rental"],
    "其他": [],
}

DEFAULT_CATEGORY = "其他"

def get_transport_subcategory(text, fn):
    """获取交通出行的细分类别"""
    t = text.lower()
    fn_lower = fn.lower()
    
    # 按优先级检查子分类
    for subcat, keywords in TRANSPORT_SUBCATEGORIES.items():
        if subcat == "其他":
            continue
        for kw in keywords:
            if kw.lower() in t or kw.lower() in fn_lower:
                return subcat
    
    # 默认其他
    return "其他"

def extract_pdf_text(pdf_path):
    try:
        r = subprocess.run(["pdftotext", "-enc", "UTF-8", pdf_path, "-"], capture_output=True, text=True, timeout=30)
        if r.returncode == 0 and r.stdout.strip():
            return r.stdout
    except: pass
    try:
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            img_path = os.path.join(tmpdir, "page")
            subprocess.run(["pdftoppm", "-png", "-singlepage", pdf_path, img_path], capture_output=True, timeout=30)
            for suffix in [".1.png", ".png"]:
                if os.path.exists(img_path + suffix):
                    r = subprocess.run(["tesseract", img_path + suffix, "stdout", "-l", "chi_sim+eng", "--psm", "6"], capture_output=True, text=True, timeout=60)
                    if r.returncode == 0 and r.stdout.strip():
                        return r.stdout
    except: pass
    return ""

def detect_currency(text):
    if re.search(r"US\$|USD|Dollar", text, re.I): return "USD"
    if re.search(r"HK\$|HKD|港币|港元", text, re.I): return "HKD"
    if re.search(r"JP¥|JPY|日元", text, re.I): return "JPY"
    if re.search(r"GBP|£", text, re.I): return "GBP"
    if re.search(r"EUR|€", text, re.I): return "EUR"
    return "CNY"

def extract_amount(text):
    currency = detect_currency(text)
    sym_map = {"CNY": "￥", "USD": "$", "HKD": "HK$", "JPY": "￥", "GBP": "£", "EUR": "€"}
    sym = sym_map.get(currency, "¥")
    # 1. 小写（含税）- 支持（ 小 写 ）格式（中间有空格）
    for p in [r"小\s*写(?:\s*[）\)])?\s*[￥¥$]?\s*([0-9,]+\.?\d*)", r"小\s*写[：:]\s*[￥¥$]?\s*([0-9,]+\.?\d*)"]:
        m = re.search(p, text, re.DOTALL)
        if m:
            try:
                v = float(m.group(1).replace(",", ""))
                if 0 < v < 1000000: return v, currency
            except: pass
    # 2. 价税合计
    for p in [r"价税合计[^\n]*?\s*[" + sym + r"]\s*([0-9,]+\.?\d*)", r"含税金额[：:]?\s*[" + sym + r"]?\s*([0-9,]+\.?\d*)"]:
        m = re.search(p, text, re.DOTALL)
        if m:
            try:
                v = float(m.group(1).replace(",", ""))
                if 0 < v < 1000000: return v, currency
            except: pass
    # 3. 总计/合计
    if currency == "USD":
        patterns = [r"总计[^\$]*\$([0-9,]+\.?\d*)", r"\$([0-9,]+\.?\d*)"]
    else:
        patterns = [r"总计[：:]*\s*[" + sym + r"]([0-9,]+\.?\d*)", r"合计[：:]*\s*[" + sym + r"]([0-9,]+\.?\d*)"]
    for p in patterns:
        m = re.search(p, text, re.DOTALL)
        if m:
            try:
                v = float(m.group(1).replace(",", ""))
                if 0 < v < 100000: return v, currency
            except: pass
    return 0.0, currency

def extract_invoice_sig(text):
    sig = []
    for p in [r"发票代码[：:]?\s*([0-9A-Z]{10,20})", r"发票号[码]?[：:]?\s*([0-9A-Z]{8,20})"]:
        m = re.search(p, text, re.I)
        if m: sig.append(m.group(1).strip()); break
    else: sig.append("")
    m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text)
    if m: sig.append(f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}")
    else: sig.append("")
    m = re.search(r"[（(]\s*小\s*写(?:\s*[）\)])?\s*[￥¥$]?\s*([0-9,]+\.?\d*)", text, re.DOTALL)
    if m:
        try: sig.append(f"{float(m.group(1).replace(',', '')):.2f}")
        except: sig.append("")
    else: sig.append("")
    return tuple(sig)

def is_no_invoice_code_file(fn, sig):
    """判断是否是『无发票号』文件（Uber等），这类文件不用发票签名去重，只用MD5去重"""
    fn_lower = fn.lower()
    # Uber / 境外网约车文件
    if any(kw in fn_lower for kw in ['uber', 'lyft', 'grab']):
        return True
    # 没有发票代码的文件（sig[0]为空，且是境外票据特征）
    if not sig[0]:  # 没有发票代码
        t = sig[1].lower() if sig[1] else ''  # date text
        # USD/HKD 等外币，且没有发票号 → 大概率是境外网约车
        if sig[2] and float(sig[2].replace(',','')) > 0:  # amount exists
            return True
    return False

def extract_summary(text, fn, amount, currency):
    """从PDF文本提取『项目名称 + 金额』干净摘要"""
    cs = {"CNY": "¥", "USD": "US$", "HKD": "HK$"}
    sym = cs.get(currency, "¥")
    amt_str = f"{sym}{amount:.2f}"
    fn_lower = fn.lower()
    
    # Uber / 境外网约车
    if 'uber' in fn_lower:
        # 尝试从文本提取 Uber 合计金额
        m = re.search(r"总计\s*US\$\s*([\d,]+\.?\d*)", text)
        if m:
            try:
                v = float(m.group(1).replace(",", ""))
                return f"Uber 运输服务 US${v:.2f}"
            except: pass
        return f"Uber 运输服务 {amt_str}"
    
    # 铁路电子客票
    if '铁路' in text or '火车票' in fn_lower:
        m = re.search(r"电子客票.*?票价:￥?([\d,]+\.?\d*)", text, re.DOTALL)
        if m:
            try:
                v = float(m.group(1).replace(",", ""))
                return f"铁路电子客票 ¥{v:.2f}"
            except: pass
        # 高铁
        m = re.search(r"高铁.*?票价:￥?([\d,]+\.?\d*)", text, re.DOTALL)
        if m:
            try:
                v = float(m.group(1).replace(",", ""))
                return f"高铁客票 ¥{v:.2f}"
            except: pass
        return f"铁路客票 {amt_str}"
    
    # 机票/航空
    if any(k in text for k in ["航空", "机票", "flight", "airline"]):
        m = re.search(r"票价:￥?([\d,]+\.?\d*)", text)
        if m:
            try:
                v = float(m.group(1).replace(",", ""))
                return f"航空客票 ¥{v:.2f}"
            except: pass
        return f"航空客票 {amt_str}"
    
    # 收费公路/ETC
    if any(k in text for k in ["收费公路", "通行费", "etc"]):
        return f"收费公路通行费 {amt_str}"
    
    # 汽油
    if '汽油' in text or '加油' in fn_lower:
        m = re.search(r"\*汽油\*([^\n]+?)\s*\n", text)
        if m:
            name = re.sub(r'\s+', '', m.group(1))
            return f"汽油/{name[:10]} {amt_str}"
        return f"汽油费 {amt_str}"
    
    # 停车费
    if '停车' in text or '停车费' in fn_lower:
        return f"停车费 {amt_str}"
    
    # 充电
    if '充电' in text or '充电桩' in fn_lower:
        return f"充电电费 {amt_str}"
    
    # 滴滴/高德网约车
    if any(k in fn_lower for k in ['滴滴', '高德', '阳光出行']) and '电子发票' in fn_lower:
        return f"运输服务 {amt_str}"
    
    # 酒店
    if any(k in text for k in ["入住", "Hotel", "住宿费", "房费"]):
        return f"酒店住宿费 {amt_str}"
    
    # 餐饮
    if any(k in text for k in ["餐费", "餐饮服务", "餐饮"]):
        return f"餐饮费 {amt_str}"
    
    # 船票
    if '船票' in fn_lower or '海天码头' in text:
        return f"船票 {amt_str}"
    
    # 办公用品 / 立创
    if '办公' in fn_lower or '立创' in fn_lower:
        m = re.search(r"项目名称\s+\*([^*\n]+)", text)
        if m:
            name = re.sub(r'\s+', '', m.group(1))
            return f"{name[:15]} {amt_str}"
        return f"办公用品 {amt_str}"
    
    # 项目名称提取（通用）
    m = re.search(r"项目名称\s+\*([^*\n]+)", text)
    if m:
        name = re.sub(r'\s+', '', m.group(1))
        return f"{name[:20]} {amt_str}"
    
    # 发票代码作名称
    m = re.search(r"发票代码[：:]?\s*([0-9A-Z]{10,20})", text)
    if m:
        return f"发票_{m.group(1)[-8:]} {amt_str}"
    
    return amt_str

def classify(text, fn):
    # 优先使用高级OCR分类
    if ADVANCED_OCR:
        main_cat, subcat = ocr_classify(text, fn)
        if main_cat != '其他':
            return main_cat
    # 备用本地分类
    t = text.lower()
    fn_lower = fn.lower()
    for cat, kws in CATEGORIES.items():
        for kw in kws:
            if kw.lower() in t or kw.lower() in fn_lower:
                return cat
    return DEFAULT_CATEGORY

def is_itinerary_not_invoice(fn, text):
    """判断是否是国内人民币行程单（而非正式发票），不纳入汇总表
    行程单关键词：行程报销单、电子行程单、行程单
    - 外币（USD/HKD等）行程单 → 保留（国外出差需要）
    - 电子发票 → 保留
    - 国内人民币行程单 → 排除"""
    fn_lower = fn.lower()
    t = text.lower()
    itinerary_keywords = ['行程报销单', '电子行程单', '行程单']
    is_itinerary = any(kw in fn_lower for kw in itinerary_keywords)
    # 如果文件名含『电子发票』，则不是行程单，直接保留
    if '电子发票' in fn_lower or '电子普通发票' in fn_lower:
        return False
    # 如果是外币行程单，保留（国外出差）
    if is_itinerary:
        # 通过文件名检测外币
        if re.search(r'usd|\$|hkd|港币|美元', fn_lower):
            return False
        # 通过PDF文本检测外币
        if re.search(r'US\$|USD|Dollar|HK\$|HKD', text, re.I):
            return False
        # 国内人民币行程单 → 排除
        return True
    return False

def scan_and_extract(fp):
    fn = os.path.basename(fp)
    text = extract_pdf_text(fp) if fp.lower().endswith(".pdf") else ""
    
    # 优先使用高级OCR提取金额
    if ADVANCED_OCR:
        amount, currency = ocr_extract_amount(text)
        date = ocr_extract_date(text) or ''
        invoice_number = ocr_extract_invoice_number(text) or ''
    else:
        amount, currency = extract_amount(text)
        date = ''
        invoice_number = ''
    sig = extract_invoice_sig(text) if amount > 0 else ("", "", "")
    if not date:
        m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text)
        date = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}" if m else datetime.now().strftime("%Y-%m-%d")
    return {"path": fp, "name": fn, "category": classify(text, fn), "amount": amount, "currency": currency, "date": date, "text": text, "sig": sig}

def scan_dir(d):
    r = []
    for root, _, files in os.walk(d):
        for f in files:
            if Path(f).suffix.lower() in [".pdf", ".jpg", ".jpeg", ".png", ".gif"]:
                r.append(os.path.join(root, f))
    return r

def gen_excel(invoices, out_path):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票汇总"
    hf = Font(bold=True, color="FFFFFF")
    hl = PatternFill(start_color="1E3A5F", fill_type="solid")
    cs = {"CNY": "¥", "USD": "US$", "HKD": "HK$", "JPY": "JP¥", "GBP": "£", "EUR": "€"}
    headers = ["序号", "文件名", "类别", "日期", "金额", "货币", "PDF摘要"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font, c.fill, c.alignment = hf, hl, Alignment(horizontal="center")
    invoices.sort(key=lambda x: (x["category"], x["date"]))
    ct = defaultdict(lambda: defaultdict(float))
    crt = defaultdict(float)
    for i, inv in enumerate(invoices, 1):
        r = i + 1
        ws.cell(r, 1, i)
        ws.cell(r, 2, inv["name"])
        ws.cell(r, 3, inv["category"])
        ws.cell(r, 4, inv["date"])
        ws.cell(r, 5, float(inv["amount"]))
        ws.cell(r, 6, inv["currency"])
        ws.cell(r, 7, extract_summary(inv.get("text", ""), inv["name"], inv["amount"], inv["currency"]))
        ct[inv["category"]][inv["currency"]] += float(inv["amount"])
        crt[inv["currency"]] += float(inv["amount"])
    tr = len(invoices) + 2
    ws.cell(tr, 2, "合计 (CNY)")
    ws.cell(tr, 2).font = Font(bold=True)
    ws.cell(tr, 5, crt.get("CNY", 0))
    ws.cell(tr, 5).font = Font(bold=True)
    ws.cell(tr, 6, "CNY")
    if "USD" in crt:
        ws.cell(tr + 1, 2, "合计 (USD)")
        ws.cell(tr + 1, 2).font = Font(bold=True)
        ws.cell(tr + 1, 5, crt.get("USD", 0))
        ws.cell(tr + 1, 5).font = Font(bold=True)
        ws.cell(tr + 1, 6, "USD")
    cr = tr + 2
    ws.cell(cr, 2, "分类汇总")
    ws.cell(cr, 2).font = Font(bold=True)
    cr += 1
    for cat in sorted(ct.keys()):
        for curr in sorted(ct[cat].keys()):
            ws.cell(cr, 2, cat)
            ws.cell(cr, 6, curr)
            ws.cell(cr, 5, ct[cat][curr])
            cr += 1
    cr += 1
    ws.cell(cr, 2, "货币总汇")
    ws.cell(cr, 2).font = Font(bold=True)
    cr += 1
    for curr in sorted(crt.keys()):
        ws.cell(cr, 2, f"合计 ({curr})")
        ws.cell(cr, 5, f"{cs.get(curr, curr)}{crt[curr]:.2f}")
        cr += 1
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 35
    wb.save(out_path)
    return crt

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", "-s", required=True)
    ap.add_argument("--output", "-o", required=True)
    ap.add_argument("--month", "-m", required=True)
    ap.add_argument("--email-dir", "-e")
    args = ap.parse_args()
    files = scan_dir(args.source)
    if args.email_dir and Path(args.email_dir).exists():
        files.extend(scan_dir(args.email_dir))
    print(f"共 {len(files)} 个文件")
    base = Path(args.output) / f"{args.month}_报销发票"
    for c in set(CATEGORIES.keys()) | {DEFAULT_CATEGORY}:
        (base / c).mkdir(parents=True, exist_ok=True)
    # 创建交通出行子目录
    for subcat in TRANSPORT_SUBCATEGORIES.keys():
        (base / "交通出行" / subcat).mkdir(parents=True, exist_ok=True)
    invoices = []
    seen_sigs = set()
    seen_content_hashes = {}  # hash -> filename, for cross-file dedup
    dup_count = 0
    for i, fp in enumerate(files, 1):
        print(f"\n[{i}/{len(files)}] {os.path.basename(fp)}")
        
        # 计算文件内容哈希
        import hashlib
        with open(fp, 'rb') as f:
            content_hash = hashlib.md5(f.read()).hexdigest()
        
        info = scan_and_extract(fp)
        
        # 跳过行程单（行程报销单/电子行程单），只保留发票
        if is_itinerary_not_invoice(info["name"], info.get("text", "")):
            print(f"  ⊝ 行程单（不纳入汇总表）")
            dup_count += 1
            continue
        
        sig_key = "|".join(str(x) for x in info["sig"])
        
        # 决定用哪种去重方式：
        # - 有发票代码的正规中国发票 → 用签名去重（避免同一张发票多次扫描）
        # - 无发票号的境外文件（Uber等）→ 用 MD5 内容哈希去重（不同行程即使金额相同也保留）
        use_sig_dedup = bool(sig_key and sig_key != "||" and info["amount"] > 0 and not is_no_invoice_code_file(info["name"], info["sig"]))
        
        if use_sig_dedup:
            # 有发票代码 → 用签名去重
            if sig_key in seen_sigs:
                print(f"  重复发票（{info['sig'][0][:15] if info['sig'][0] else 'N/A'} | {info['sig'][1]} | {info['sig'][2]}）")
                dup_count += 1
                continue
            seen_sigs.add(sig_key)
        
        # 检查内容哈希去重（文件名不同但内容相同）
        if content_hash in seen_content_hashes:
            dup_desc = f"sig去重({info['sig'][0][:12] if use_sig_dedup and info['sig'][0] else 'N/A'})" if use_sig_dedup else "MD5去重"
            print(f"  ⊝ 内容完全相同（已存在为: {seen_content_hashes[content_hash]}）")
            dup_count += 1
            continue
        seen_content_hashes[content_hash] = os.path.basename(fp)
        
        print(f"  {info['category']} | {info['currency']}{info['amount']:.2f}")
        
        # 交通出行使用细分类
        category = info["category"]
        if category == "交通出行":
            subcat = get_transport_subcategory(info.get("text", ""), info["name"])
            category = f"交通出行/{subcat}"
            info["category"] = category
            print(f"    → {subcat}")
        
        ext = Path(info["name"]).suffix
        # 保持原始文件名，但如果同名文件已存在且内容不同，则加后缀区分
        new_name = info["name"]
        dest = base / category / new_name
        if dest.exists():
            # 同名文件已存在，比较内容哈希
            import hashlib
            with open(fp, 'rb') as f1:
                new_hash = hashlib.md5(f1.read()).hexdigest()
            with open(dest, 'rb') as f2:
                existing_hash = hashlib.md5(f2.read()).hexdigest()
            if new_hash != existing_hash:
                # 内容不同，加后缀
                stem = Path(info["name"]).stem
                suffix = 2
                while True:
                    new_name = f"{stem}_{suffix}{ext}"
                    dest = base / category / new_name
                    if not dest.exists():
                        break
                    suffix += 1
                print(f"    ⚠️ 同名不同内容，添加后缀: {new_name}")
        try:
            import shutil
            shutil.copy2(fp, dest)
        except: pass
        invoices.append(info)
    print(f"\n完成！处理 {len(files)} 个文件，去重 {dup_count} 个")
    out_excel = base / "发票汇总表.xlsx"
    totals = gen_excel(invoices, str(out_excel))
    cs = {"CNY": "¥", "USD": "US$", "HKD": "HK$"}
    for curr in sorted(totals.keys()):
        print(f"  {curr}: {cs.get(curr, curr)}{totals[curr]:.2f}")
    print(f"\n汇总表: {out_excel}")

if __name__ == "__main__":
    main()
