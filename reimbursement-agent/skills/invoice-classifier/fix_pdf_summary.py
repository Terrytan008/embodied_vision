#!/usr/bin/env python3
"""修复发票汇总表G列：PDF摘要 = 干净的「项目名称 ¥金额」"""
import openpyxl, subprocess, re, os, glob

REIMB_DIR = "/Users/terrytan/报销发票/2026-03_报销发票"
EXCEL     = "/Users/terrytan/openclaw/workspace/发票汇总表_2026-03.xlsx"

def extract_pdf_text(pdf_path):
    try:
        r = subprocess.run(["pdftotext", "-enc", "UTF-8", pdf_path, "-"],
                            capture_output=True, text=True, timeout=30)
        return r.stdout
    except Exception as e:
        return "[ERROR: %s]" % e

_SMALL = u"[零壹贰叁肆伍陆柒捌玖]+"
_BIG   = u"[零壹贰叁肆伍陆柒捌玖拾佰仟万圆角分整]+"

def parse_chinese_amount(text):
    """从PDF文本中提取中文大写金额"""
    # 查找价税合计后面跟着的中文数字
    patterns = [
        u"价税合计[^]{0,30}?" + _BIG,          # 价税合计（大写）...壹佰圆整
        u"价税合计[^]{0,30}?" + _SMALL + u"[圆元]" + _SMALL + u"[角分]",  # 叁佰圆整
        _BIG,
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            cnum = m.group(0)
            # 简单解析
            total = 0.0
            yuan = 0.0; jiao = 0.0; fen = 0.0
            num_map = {u'零':0,u'一':1,u'壹':1,u'二':2,u'贰':2,u'三':3,u'叁':3,
                       u'四':4,u'肆':4,u'五':5,u'伍':5,u'六':6,u'陆':6,
                       u'七':7,u'柒':7,u'八':8,u'捌':8,u'九':9,u'玖':9}
            unit_map = {u'拾':10,u'佰':100,u'仟':1000,u'万':10000}
            tmp = 0.0; last_unit = 1
            i = 0
            while i < len(cnum):
                ch = cnum[i]
                if ch in num_map:
                    v = num_map[ch]
                    if i+1 < len(cnum):
                        nxt = cnum[i+1]
                        if nxt in unit_map:
                            tmp += v * unit_map[nxt]
                            i += 2; continue
                        elif nxt == u'圆' or nxt == u'元':
                            tmp += v
                            i += 2; continue
                        elif nxt == u'角':
                            jiao = v; i += 2; continue
                        elif nxt == u'分':
                            fen = v; i += 2; continue
                    tmp += v
                elif ch == u'整':
                    yuan += tmp; tmp = 0
                elif ch == u'圆' or ch == u'元':
                    pass
                i += 1
            total = yuan + tmp + jiao * 0.1 + fen * 0.01
            if total > 0:
                return total
    return None

def extract_summary(text, filename=""):
    """从PDF文本中提取「项目名称 + 金额」摘要"""
    # 小写金额的正则（兼容全角/半角括号、空格、￥/¥）
    def money(pat=None):
        if pat:
            m = re.search(pat, text)
        else:
            m = re.search(u"[(（]\s*小\s*写\s*[)）]\s*[￥¥]?\s*([\d,]+\\.?\d*)", text)
        if m:
            return float(m.group(1).replace(",",""))
        return None

    # 1. 铁路电子客票
    m = re.search(u"电子发票（铁路电子客票）.*?票价:￥?([\d,]+\\.?\d*)", text, re.DOTALL)
    if m:
        return u"铁路电子客票 ¥%.2f" % float(m.group(1).replace(",",""))

    # 2. 收费公路通行费（ETC）
    m = re.search(u"收费公路通行费电子票据.*?交易金额\s*¥?([\d,]+\\.?\d*)", text, re.DOTALL)
    if m:
        return u"收费公路通行费 ¥%.2f" % float(m.group(1).replace(",",""))

    # 3. 汽油
    m = re.search(u"\\*汽油\\*([^\n]+)", text)
    if m:
        name = re.sub(u"\s+", u"", m.group(1))
        am = money()
        if am:
            return u"汽油/%s ¥%.2f" % (name, am)

    # 4. 停车费
    if u"*经营租赁*停车费" in text:
        am = money()
        if am:
            return u"停车费 ¥%.2f" % am

    # 5. 滴滴电子发票（嵌入中文发票，肆拾叁圆整）
    if re.search(u"旅客运输服务", text) and re.search(u"肆拾叁圆整", text):
        return u"运输服务 ¥43.00"

    # 6. 滴滴电子发票（中文大写金额，如壹佰贰拾壹圆捌角整）
    if u"滴滴" in text and (u"圆" in text or u"角" in text):
        amt = parse_chinese_amount(text)
        if amt and amt > 0:
            return (u"滴滴出行 ¥%.2f") % amt
        m = re.search(u"合计\s*([\d,]+\\.?\d*)\s*元", text)
        if m:
            return (u"滴滴出行 ¥%.2f") % float(m.group(1).replace(",",""))

    # 7. 滴滴行程报销单
    if (u"滴滴出行" in text or u"DIDI" in text.upper()) and u"合计" in text:
        m = re.search(u"合计\s*([\d,]+\\.?\d*)\s*元", text)
        if m:
            return (u"滴滴出行 ¥%.2f") % float(m.group(1).replace(",",""))

    # 8. 携程电子发票（经纪代理服务-代订机票/酒店）
    if u"经纪代理服务" in text and u"代订" in text:
        am = money()
        if am:
            return u"代订服务费 ¥%.2f" % am

    # 9. Uber 邮件（含电子发票）
    if u"Uber" in text or u"优步" in text:
        cn_m = re.search(u"项目名称\s+\\*([^\n]+?)\s*\n[^\n]*\n[^\n]*\n[^\n]*\n.*?[(（]\s*小\s*写\s*[)）]\s*[￥¥]?\s*([\d,]+\\.?\d*)", text, re.DOTALL)
        if cn_m:
            name = re.sub(u"\s+", u"", cn_m.group(1))
            return u"%s ¥%.2f" % (name, float(cn_m.group(2).replace(",","")))
        m = re.search(u"总计\s*US\\$\s*([\d.]+)", text)
        if m:
            return u"Uber ¥%.2f(US)" % float(m.group(1))

    # 10. 高德打车电子发票（含嵌入电子发票格式）
    if u"高德" in text and u"打车" in text:
        if re.search(u"项目名称\s+\\*运输服务\\*客运服务费", text):
            am = money()
            if am:
                return u"运输服务 ¥%.2f" % am
        m = re.search(u"合计\s*([\d,]+\\.?\d*)\s*元", text)
        if m:
            return u"网约车 ¥%.2f" % float(m.group(1).replace(",",""))

    # 11. 阳光出行（含行程报销单）
    if u"阳光出行" in text:
        m = re.search(u"票价:￥?([\d,]+\\.?\d*)", text)
        if m:
            return u"客运服务费 ¥%.2f" % float(m.group(1).replace(",",""))
        am = money()
        if am:
            return u"运输服务 ¥%.2f" % am
        m = re.search(u"合计\s*([\d,]+\\.?\d*)\s*元", text)
        if m:
            return u"阳光出行 ¥%.2f" % float(m.group(1).replace(",",""))

    # 12. 代驾
    if u"代驾" in text:
        m = re.search(u"合计\s*([\d,]+\\.?\d*)\s*元", text)
        if m:
            return u"代驾服务费 ¥%.2f" % float(m.group(1).replace(",",""))
        am = money()
        if am:
            return u"代驾服务费 ¥%.2f" % am

    # 13. 旅游服务
    if u"旅游服务" in text:
        am = money()
        if am:
            pm = re.search(u"项目名称\s+\\*([^\n]+)", text)
            if pm:
                name = re.sub(u"\s+", u"", pm.group(1))
                return u"%s ¥%.2f" % (name, am)
            return u"旅游服务 ¥%.2f" % am

    # 14. 供电/充电
    if any(k in text for k in [u"充电", u"供电"]):
        m = re.search(u"项目名称\s+\\*([^\n]+)", text)
        am = money()
        if m and am:
            name = re.sub(u"\s+", u"", m.group(1))
            return u"%s ¥%.2f" % (name, am)

    # 15. 酒店水单（非电子发票格式）
    if u"宾客水单" in text or u"酒店名称" in text:
        m = re.search(u"总金额\s*([\d,]+\\.?\d*)", text)
        if m:
            return u"酒店住宿费 ¥%.2f" % float(m.group(1).replace(",",""))

    # 16. 餐饮
    if u"餐饮服务" in text or u"*餐饮" in text:
        am = money()
        if am:
            return u"餐饮费 ¥%.2f" % am

    # 17. 普通电子发票（通用）
    m = re.search(u"项目名称\s+\\*([^\n]+)", text)
    if m:
        am = money()
        if am:
            name = re.sub(u"\s+", u"", m.group(1))
            return u"%s ¥%.2f" % (name, am)

    # 18. USD金额（从文件名推断）
    if u"USD" in text.upper() or u"usd" in text:
        m = re.search(u"([\d,]+\\.?\d*)\s*USD", text, re.IGNORECASE)
        if m:
            return u"网约车 ¥%.2f(US)" % float(m.group(1).replace(",",""))

    # 19. 入住证明（英文）
    if u"Proof of Check-in" in text:
        m = re.search(u"(?:Total|Amount|Sum)\s*[:\s]*\\$?([\d,]+\\.?\d*)", text, re.IGNORECASE)
        if m:
            return u"酒店住宿费 ¥%.2f" % float(m.group(1).replace(",",""))


    # 18b. 入住凭证（中文）
    m = re.search(u"¥([\d,]+\.?\d*)\s*已支付", text)
    if m:
        return u"酒店住宿费 ¥%.2f" % float(m.group(1).replace(",",""))

    # 18c. Proof of Check-in (English)
    m = re.search(u"CNY\s*([\d,]+\.?\d*)\s*Paid", text, re.IGNORECASE)
    if m:
        return u"酒店住宿费 ¥%.2f" % float(m.group(1).replace(",",""))

    # 18d. USD amount from filename
    if filename:
        m = re.search(u"([\d,]+\.?\d*)\s*USD", filename, re.IGNORECASE)
        if m:
            return u"网约车 ¥%.2f(US)" % float(m.group(1).replace(",",""))

    return "[未能解析]"

wb = openpyxl.load_workbook(EXCEL)
ws = wb.active

print("处理表: %s, 共 %d 行数据" % (ws.title, ws.max_row-1))

for row_idx in range(2, ws.max_row + 1):
    filename = ws.cell(row_idx, 2).value

    if not filename:
        continue

    if not filename.lower().endswith('.pdf'):
        continue

    pdf_candidates = glob.glob(
        os.path.join(REIMB_DIR, "**", filename),
        recursive=True
    )
    pdf_candidates = [p for p in pdf_candidates if p.endswith(".pdf")]

    if not pdf_candidates:
        ws.cell(row_idx, 7, "[文件未找到]")
        print("  Row %d: \u274c 未找到 %s" % (row_idx, filename))
        continue

    pdf_path = pdf_candidates[0]
    text = extract_pdf_text(pdf_path)
    summary = extract_summary(text, filename)
    ws.cell(row_idx, 7, summary)
    status = "\u2705" if "\u672a\u80fd\u89e3\u6790" not in summary else "\u26a0\ufe0f"
    print("  Row %d: %s %s \u2192 %s" % (row_idx, status, filename, summary))

wb.save(EXCEL)
print("\n\u5b8c\u6210\uff01\u5df2\u4fdd\u5b58: %s" % EXCEL)
